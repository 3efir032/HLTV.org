from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

processes_xlsx = []
# Создает(открывает) файл static.xlsx и создает страницу data + устанавливает ширину столбцов А и В
def document():
    fn = 'Event/static.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30
    wb.save('Event/static.xlsx')
    return ws, wb

# Настройка отображения таблицы
def setting_up_a_table():  # настройка таблицы
    ws, wb = document()

    # Нумерация строк от 1 до 500
    def line_numbering():
        count = 1
        for x in range(6, 507, 2):
            ws.merge_cells(f'A{str(x)}:A{str(x + 1)}')  # Объединения ячеек
            ws['A' + str(x)] = count
            ws['A' + str(x)].fill = PatternFill('solid', fgColor="32CD32")
            count += 1
            for i in ['C', 'D', 'E', 'F', 'G', 'H']:
                ws.merge_cells(f'{i}{str(x)}:{i}{str(x + 1)}')  # Объединения ячеек
        #print('line_numbering - ok')
        processes_xlsx.append('Нумерация строк .xlsx')
        wb.save('Event/static.xlsx')

    # Нумерует столбы от 2 до 80, без 1 и 16
    def column_numbering():
        num = 9
        for i in range(2, 100):
            if num > 16:
                i += 1
            ws.cell(row=5, column=num, value=i)
            ws.column_dimensions[get_column_letter(num)].width = 3  # Размер столбцов 3мм
            num += 1
        #print('column_numbering - ok')
        processes_xlsx.append('Нумерация столбцов .xlsx')
        wb.save('Event/static.xlsx')

    # Цвет колонок от 2 до 80
    def color_columns():
        for row in ws.iter_rows(min_row=5, max_col=106, max_row=5):
            for cell in row:
                if cell.value == 2 or cell.value == 16 or cell.value == 31:
                    cell.fill = PatternFill('solid', fgColor="FFA500")
                    cell.font = Font(name='Calibri', size=13)
                else:
                    cell.fill = PatternFill('solid', fgColor="32CD32")
                    cell.font = Font(name='Calibri', size=13)
        #print('color_columns - ok')
        processes_xlsx.append('Создание цвета колонок .xlsx')
        wb.save('Event/static.xlsx')

    # Выделяет 2, 16 и 30 колонку - серым цветом
    def color_grey_columns():
        for c in range(6, 508):
            num_i = ws['I' + str(c)]
            num_w = ws['V' + str(c)]
            num_ak = ws['AK' + str(c)]
            num_i.fill = PatternFill('solid', fgColor="808080")
            num_w.fill = PatternFill('solid', fgColor="808080")
            num_ak.fill = PatternFill('solid', fgColor="808080")
        #print('color_grey_columns - ok')
        processes_xlsx.append('Выделение колонок .xlsx')
        wb.save('Event/static.xlsx')

    # Ширина ячеек статистики игр
    def static_setting():
        num = 3
        for i in range(2, 9):
            ws.column_dimensions[get_column_letter(num)].width = 3  # Размер столбцов 3мм
            num += 1

        column_c = ws['C5']
        column_c.value = 'РA'
        column_c.fill = PatternFill('solid', fgColor="FFA500")
        column_d = ws['D5']
        column_d.value = 'УВ'
        column_d.fill = PatternFill('solid', fgColor="A52A2A")
        column_d = ws['E5']
        column_d.value = 'В'
        column_d.fill = PatternFill('solid', fgColor="808080")
        column_d = ws['F5']
        column_d.value = 'Д'
        column_d.fill = PatternFill('solid', fgColor="800080")
        column_d = ws['G5']
        column_d.value = 'С'
        column_d.fill = PatternFill('solid', fgColor="2F4F4F")
        column_d = ws['H5']
        column_d.value = 'П'
        column_d.fill = PatternFill('solid', fgColor="000000")

        for i in range(6, 508, 2):
            ws[f'C{i}'] = f'=COUNTA(I{i}:DB{i + 1})'
            ws[f'D{i}'] = f'=COUNTIF(I{i}:DB{i + 1};1) + COUNTIF(I{i}:DB{i + 1};2)'
            ws[f'E{i}'] = f'=COUNTIF(I{i}:DB{i + 1};3)'
            ws[f'F{i}'] = f'=COUNTIF(I{i}:DB{i + 1};4)'
            ws[f'G{i}'] = f'=COUNTIF(I{i}:DB{i + 1};5)'
            ws[f'H{i}'] = f'=C{i}-D{i}-E{i}-F{i}-G{i}'
        #print('static_setting - ok')
        processes_xlsx.append('Редактирование ширины .xlsx')
        wb.save('Event/static.xlsx')

    # Создание строк основной статистики и подсчет
    def statistics_fields():
        ws.merge_cells('D3:G3')
        ws.merge_cells('H3:J3')
        ws['D3'] = 'Раундов'
        ws['H3'] = '=sum(C6:C506)'

        ws.merge_cells('D4:G4')
        ws.merge_cells('H4:J4')
        ws['D4'] = 'Убийство всех'
        ws['H4'] = '=sum(D6:D506)'

        ws.merge_cells('K3:N3')
        ws.merge_cells('O3:Q3')
        ws['K3'] = 'Взрыв'
        ws['O3'] = '=sum(E6:E506)'

        ws.merge_cells('K4:N4')
        ws.merge_cells('O4:Q4')
        ws['K4'] = 'Дефуз'
        ws['O4'] = '=sum(F6:F506)'

        ws.merge_cells('R3:U3')
        ws.merge_cells('V3:X3')
        ws['R3'] = 'Сейв'
        ws['V3'] = '=sum(G6:G506)'

        ws.merge_cells('R4:U4')
        ws.merge_cells('V4:X4')
        ws['R4'] = 'Карт'
        ws['V4'] = '=COUNTA(B6:B506)/2'

        #print('statistics_fields - ok')
        processes_xlsx.append('Создание статистики подсчета .xlsx')
        wb.save('Event/static.xlsx')

    # Выравниваем заполненные ячейки, только имеющие значения
    def alignment_font():
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False,
                                               shrink_to_fit=True)
                    cell.font = Font(name='Calibri', size=11)  # размер и стиль, шрифта документа
        #print('alignment_font - ok')
        processes_xlsx.append('Выравнивание ячеек .xlsx')
        wb.save('Event/static.xlsx')

    # Рисует границы игр
    def border_column():
        letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                   'U',
                   'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL',
                   'AM',
                   'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD',
                   'BE',
                   'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV',
                   'BW',
                   'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN',
                   'CO',
                   'CP', 'CQ', 'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB'
                   ]

        thins = Side(border_style="thin", color="000000")
        for x in range(7, 508, 2):
            for letter in letters:
                cell = ws[f'{letter}{x}']
                cell.border = Border(bottom=thins)
        #print('border_column - ok')
        processes_xlsx.append('Создание границ .xlsx')
        wb.save('Event/static.xlsx')


    return line_numbering(), column_numbering(), color_columns(), color_grey_columns(), static_setting(), statistics_fields(), border_column(), alignment_font()
