from PyQt5.QtCore import QThread
from PyQt5.QtWidgets import QApplication, QFileDialog, QMessageBox
from PyQt5.QtGui import QIcon
from hltv import *
from openpyxl.workbook import Workbook
from mainpage import MatchCS
from gamepages import HLTV
from xlms import setting_up_a_table
import time
import sys
import os


class ParsingMaps(QThread):
    def __init__(self, mainwindow, parent=None):
        super().__init__()
        self.mainwindow = mainwindow

    def run(self):
        # Создает файл 'static.xlsx' и страницу 'data'
        def creating_a_file():
            wb = Workbook()
            ws = wb.active
            ws.title = 'data'
            wb.save('Event\static.xlsx')

            # Модюль отбражения загрузки creating_a_file() - QT
            self.mainwindow.loading_info.setText('Создание файла -.xlsx')
            for i in range(0, 13):
                time.sleep(0.5)
                self.mainwindow.progressBar.setValue(i)
            time.sleep(0.3)
            self.mainwindow.loading_info.setText('Файл .xlsx - Создан')
            self.mainwindow.progressBar.setValue(13)
            time.sleep(1.5)
            self.mainwindow.loading_info.setText('Поиск статистики чемпионата')
            self.mainwindow.progressBar.setValue(14)

            return ws, wb

        ws, wb = creating_a_file()

        # Парсит сайт и собирает результат по играм и картам
        def fi_result_info(url):

            # Запрос по статистики
            match = MatchCS(url)
            maps = match.maps  # Ссылки на карты в матчах
            name_map = []  # Список карт
            event_name = match.eventName  # Название события
            team_result = []  # Общий результат
            match_team1 = []  # Результат 1-й команды
            match_team2 = []  # Результат 2-й команды
            self.mainwindow.loading_info.setText('Начинаю сбор статистики по картам')
            self.mainwindow.progressBar.setValue(15)
            for onemap in maps:
                time.sleep(0.3)
                hltv = HLTV(onemap)  # парсинг страниц карт
                name_map.append(hltv.nameMaps)  # Добавляет название карт в Список карт
                result_team, result_1, result_2 = hltv.methodWinRound
                if result_team != []:
                    result_team.append(team_result)
                    match_team1.append(result_1)
                    match_team2.append(result_2)

            self.mainwindow.loading_info.setText('Сохранение статистики')
            for i in range(16, 50):
                time.sleep(0.2)
                self.mainwindow.progressBar.setValue(i)


            count = 0
            for i in range(len(name_map)):
                ws["B" + str((i + count) + 6)] = name_map[i]
                ws["B" + str((i + count) + 7)] = 'https://www.hltv.org' + maps[i]
                count += 1
            ws['B3'] = event_name
            wb.save('Event\static.xlsx')

            return match, team_result, match_team1, match_team2

        # Фильтрует результат 1-й команды и добавляет в файл
        def exel_result_team1(one):

            match_team1 = one
            exls = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                    'V',
                    'W',
                    'X', 'Y', 'Z',
                    'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP',
                    'AQ',
                    'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
                    'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP',
                    'BQ',
                    'BR',
                    'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
                    'CA', 'CB', 'CC', 'CD', 'C', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ',
                    'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ']
            res = match_team1
            kill = 0
            bomb = 0
            bomb_defuse = 0
            stop_watch = 0
            count = -1
            nums = 7
            self.mainwindow.loading_info.setText('Сортирую результат 1-й команды')
            for i in range(50, 71):
                time.sleep(0.2)
                self.mainwindow.progressBar.setValue(i)

            for item in range(0, len(res)):
                count += 1
                for i in range(0, len(res[item])):
                    # item - число строк
                    # i - число содержания в списке

                    if res[item][i] == "emptyHistory":
                        res[item][i] = ''
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 't_win':
                        res[item][i] = '1'
                        kill += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'ct_win':
                        res[item][i] = '2'
                        kill += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'bomb_exploded':
                        res[item][i] = '3'
                        bomb += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'bomb_defused':
                        res[item][i] = '4'
                        bomb_defuse += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'stopwatch':
                        res[item][i] = '5'
                        stop_watch += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    else:
                        res[item][i] = None
            wb.save('Event\static.xlsx')
            self.mainwindow.loading_info.setText('Сортировка результатов 1-й команды - Завершена')
            self.mainwindow.progressBar.setValue(71)
            return kill, bomb, bomb_defuse, stop_watch

        # Фильтрует результат 2-й команды и добавляет в файл
        def exel_result_team2(one):
            match_team2 = one
            exls = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                    'V',
                    'W',
                    'X', 'Y', 'Z',
                    'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP',
                    'AQ',
                    'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
                    'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP',
                    'BQ',
                    'BR',
                    'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
                    'CA', 'CB', 'CC', 'CD', 'C', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ',
                    'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ']
            res = match_team2
            kill = 0
            bomb = 0
            bomb_defuse = 0
            stop_watch = 0
            count = 0
            nums = 7
            self.mainwindow.loading_info.setText('Сортирую результат 2-й команды')
            for i in range(72, 83):
                time.sleep(0.2)
                self.mainwindow.progressBar.setValue(i)

            for item in range(len(res)):
                count += 1
                for i in range(len(res[item])):

                    if res[item][i] == "emptyHistory":
                        res[item][i] = ''
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 't_win':
                        res[item][i] = '1'
                        kill += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'ct_win':
                        res[item][i] = '2'
                        kill += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'bomb_exploded':
                        res[item][i] = '3'
                        bomb += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'bomb_defused':
                        res[item][i] = '4'
                        bomb_defuse += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    elif res[item][i] == 'stopwatch':
                        res[item][i] = '5'
                        stop_watch += 1
                        ws[exls[i + nums] + str(6 + item + count)] = res[item][i]
                    else:
                        res[item][i] = None
            wb.save('Event\static.xlsx')
            self.mainwindow.loading_info.setText('Сортировка результатов 2-й команды - Завершена')
            self.mainwindow.progressBar.setValue(83)
            self.mainwindow.loading_info.setText('Подготовка к сохранению файла')
            self.mainwindow.progressBar.setValue(84)
            return kill, bomb, bomb_defuse, stop_watch

        def launch():
            place = self.mainwindow.event_number.text()
            error = QMessageBox()
            if place.isdigit():
                url = place
                match, team_result, match_team1, match_team2 = fi_result_info(url)
                exel_result_team1(match_team1)
                exel_result_team2(match_team2)
                setting_up_a_table()  # from xlms import setting_up_a_table

                self.mainwindow.loading_info.setText('Сохранение результатов')
                self.mainwindow.progressBar.setValue(85)

                old_file = os.path.join("Event", "static.xlsx")
                new_file = os.path.join("Event", f"{match.eventName}.xlsx")
                os.rename(old_file, new_file)

                self.mainwindow.loading_info.setText('Удаление временных файлов')
                for i in range(86, 101):
                    time.sleep(0.1)
                    self.mainwindow.progressBar.setValue(i)


                self.mainwindow.loading_info.setText('ГОТОВО')

            elif place == '':
                error.setWindowTitle('WARNING')
                error.setText(
                    '!!!ПУСТАЯ СТРОКА!!!\n\nДобавьте цифры, после знака "=".\n\nПРИМЕР:\nhttps://www.hltv.org/results?event=2062\nВ строку ввода добавь: 2062')  # Текс ошибки
                error.setIcon(QMessageBox.Warning)
                error.setStandardButtons(QMessageBox.Ok)
                error.exec_()

            else:
                error.setWindowTitle('WARNING')  # Название ошибки
                error.setText(
                    '!!!НЕКОРРЕКТНЫЙ НОМЕР СОБЫТИЯ!!!\n\nДобавьте цифры, после знака "=".\n\nПРИМЕР:\nhttps://www.hltv.org/results?event=2062\nВ строку ввода добавь: 2062')  # Текс ошибки
                error.setIcon(QMessageBox.Warning)  # Иконка ошибки
                error.setStandardButtons(QMessageBox.Ok)  # Кнопка ОК, по умолчанию закрывает ошибку
                error.exec_()  # Отображение окна

        return launch()


class Window(QtWidgets.QMainWindow):
    def __init__(self):
        super(Window, self).__init__()
        self.ui = Ui_HltvWindow()
        self.ui.setupUi(self)
        self.init_UI()

    def init_UI(self):
        self.setWindowIcon(QIcon('icon\icon.png'))
        self.ui.event_number.setPlaceholderText('EVENT')
        self.ui.loading_info.setText('Информация о загрузке')
        self.ui.pushButton_start.pressed.connect(self.start_parsing)
        self.Parsing = ParsingMaps(mainwindow=self.ui)

    def start_parsing(self):
        self.Parsing.start()


if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    application = Window()
    application.show()
    sys.exit(app.exec_())
